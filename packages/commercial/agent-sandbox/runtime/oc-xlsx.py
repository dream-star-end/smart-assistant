#!/usr/bin/env python3
# oc-xlsx — 容器内确定性 Excel(.xlsx)生成 CLI(办公助手能力,openpyxl,中文友好默认样式)。
#
# 用法(baseline skill office-spreadsheet 文档化):
#   oc-xlsx from-csv  <input.csv>  [-o out.xlsx] [--sheet 名] [--no-header]
#   oc-xlsx from-json <input.json> [-o out.xlsx] [--sheet 名]   # JSON 为对象数组
#   oc-xlsx from-md   <input.md>   [-o out.xlsx]                # markdown 表格(可多张→多 sheet)
#
# 复杂需求(公式/透视/多表关联/图表/大数据分析)直接写 Python 用 openpyxl / pandas /
# duckdb(见 office-spreadsheet skill),本 CLI 只做最高频的"结构化数据 → 规范 Excel"一击:
#   表头加粗+底纹、冻结首行、自动列宽、中文字体、数字右对齐。零许可风险(openpyxl=MIT)。
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as e:  # pragma: no cover - 依赖缺失时清晰报错
    sys.stderr.write(f"oc-xlsx: openpyxl 不可用({e});镜像应已预装 openpyxl\n")
    sys.exit(1)

TOOL = "oc-xlsx"
# 中文字体名仅作元数据(Excel/WPS 打开时选用);内容正确性不依赖它。
CJK_FONT = "Microsoft YaHei"
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name=CJK_FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=CJK_FONT, size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fail(msg: str):
    sys.stderr.write(f"{TOOL}: {msg}\n")
    sys.exit(1)


def _num(v):
    """尽量把纯数字字符串转成数值,让 Excel 识别为数字(可参与公式/图表)。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return ""
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except ValueError:
            return s
    if re.fullmatch(r"-?\d*\.\d+", s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def _write_sheet(ws, rows, header=True):
    """rows: list[list]。第一行可作表头。应用样式 + 自动列宽 + 冻结首行。"""
    if not rows:
        return
    max_w = {}
    for r_idx, row in enumerate(rows, start=1):
        is_header = header and r_idx == 1
        for c_idx, raw in enumerate(row, start=1):
            val = raw if is_header else _num(raw)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if is_header:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = BODY_FONT
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(val, (int, float)) else "left",
                    vertical="center",
                )
            # 估算列宽(中文按 2 宽计)
            text = "" if val is None else str(val)
            w = sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)
            max_w[c_idx] = max(max_w.get(c_idx, 0), w)
    for c_idx, w in max_w.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(w + 2, 8), 60)
    if header and len(rows) > 1:
        ws.freeze_panes = "A2"


def _parse_md_tables(text):
    """从 markdown 抽取所有管道表格,返回 [(标题, rows)]。标题取表格前最近的 # 行。"""
    lines = text.splitlines()
    tables = []
    i = 0
    last_heading = None
    while i < len(lines):
        line = lines[i]
        hm = re.match(r"^#{1,6}\s+(.*)$", line.strip())
        if hm:
            last_heading = hm.group(1).strip()
        # 表头行 + 分隔行(---|:--:)
        if "|" in line and i + 1 < len(lines) and re.match(r"^\s*\|?[\s:|-]+\|?\s*$", lines[i + 1]) and "-" in lines[i + 1]:
            rows = []
            j = i
            while j < len(lines) and "|" in lines[j]:
                if re.match(r"^\s*\|?[\s:|-]+\|?\s*$", lines[j]) and "-" in lines[j]:
                    j += 1
                    continue
                cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                rows.append(cells)
                j += 1
            if rows:
                tables.append((last_heading, rows))
            last_heading = None
            i = j
            continue
        i += 1
    return tables


def _out_path(inp, explicit):
    if explicit:
        return Path(explicit)
    return Path(inp).with_suffix(".xlsx")


def cmd_from_csv(args):
    inp = args["input"]
    rows = []
    with open(inp, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            rows.append(row)
    wb = Workbook()
    ws = wb.active
    ws.title = args.get("sheet") or "Sheet1"
    _write_sheet(ws, rows, header=not args.get("no_header"))
    out = _out_path(inp, args.get("output"))
    wb.save(out)
    return out, {"sheets": 1, "rows": len(rows)}


def cmd_from_json(args):
    inp = args["input"]
    data = json.loads(Path(inp).read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data or not isinstance(data[0], dict):
        fail("from-json 需要「对象数组」,如 [{\"名称\":\"甲\",\"金额\":100}, ...]")
    cols = []
    for obj in data:
        for k in obj.keys():
            if k not in cols:
                cols.append(k)
    rows = [cols]
    for obj in data:
        rows.append([obj.get(c, "") for c in cols])
    wb = Workbook()
    ws = wb.active
    ws.title = args.get("sheet") or "Sheet1"
    _write_sheet(ws, rows, header=True)
    out = _out_path(inp, args.get("output"))
    wb.save(out)
    return out, {"sheets": 1, "rows": len(data)}


def cmd_from_md(args):
    inp = args["input"]
    tables = _parse_md_tables(Path(inp).read_text(encoding="utf-8"))
    if not tables:
        fail("未在 markdown 里找到管道表格(| ... | 且第二行为 --- 分隔)")
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for idx, (title, rows) in enumerate(tables, start=1):
        name = (title or f"表{idx}")[:28]
        base = name
        n = 1
        while name in used:
            n += 1
            name = f"{base}_{n}"[:28]
        used.add(name)
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, rows, header=True)
    out = _out_path(inp, args.get("output"))
    wb.save(out)
    return out, {"sheets": len(tables)}


def parse_args(argv):
    if not argv or argv[0] in ("-h", "--help"):
        sys.stdout.write(
            "用法:\n"
            "  oc-xlsx from-csv  <input.csv>  [-o out.xlsx] [--sheet 名] [--no-header]\n"
            "  oc-xlsx from-json <input.json> [-o out.xlsx] [--sheet 名]\n"
            "  oc-xlsx from-md   <input.md>   [-o out.xlsx]\n"
        )
        sys.exit(0)
    sub = argv[0]
    rest = argv[1:]
    args = {"no_header": False}
    positional = []
    i = 0
    while i < len(rest):
        a = rest[i]
        if a in ("-o", "--output"):
            args["output"] = rest[i + 1]
            i += 2
        elif a == "--sheet":
            args["sheet"] = rest[i + 1]
            i += 2
        elif a == "--no-header":
            args["no_header"] = True
            i += 1
        else:
            positional.append(a)
            i += 1
    if not positional:
        fail(f"{sub}: 缺少输入文件")
    args["input"] = positional[0]
    return sub, args


def main():
    sub, args = parse_args(sys.argv[1:])
    handlers = {"from-csv": cmd_from_csv, "from-json": cmd_from_json, "from-md": cmd_from_md}
    if sub not in handlers:
        fail(f"未知子命令 {sub!r}(from-csv|from-json|from-md)")
    if not Path(args["input"]).is_file():
        fail(f"输入文件不存在: {args['input']}")
    out, meta = handlers[sub](args)
    sys.stdout.write(json.dumps({"output": str(out), **meta}, ensure_ascii=False) + "\n")
    sys.stdout.write(str(out) + "\n")


if __name__ == "__main__":
    main()
